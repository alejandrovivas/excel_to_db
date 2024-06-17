from flask import Flask, request
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import random
import string
from PIL import Image as PILImage
import pandas as pd
import pymysql
import logging
from io import BytesIO
import boto3
import re
from datetime import datetime
from categorias import (categorias, HOST, PASSWORD, PORT, USER, DATABASE,AWS_BUCKET_NAME, AWS_DEFAULT_REGION,
                        AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY, AWS_DEFAULT_IMAGES_PATH)

app = Flask(__name__)

# Configure logging
timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
log_filename = f'excel_{timestamp}.log'
logging.basicConfig( filename=log_filename, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s',
                     datefmt='%Y-%m-%d %H:%M:%S')

@app.route('/upload-excel', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
        return 'No file part in the request', 400

    file = request.files['file']
    if file.filename == '':
        return 'No selected file', 400

    # try:
    # Load the workbook and the specific sheet with openpyxl
    working_book = load_workbook(file.stream)
    errors = validate_empty_cells(working_book)

    # Pre validate data
    if errors:
        return f'Existe errores en los siguientes datos {errors}', 400
    else:
        # We change to False the active status of each product
        execute_sql("UPDATE producto SET activo = 0 WHERE activo = 1;")

        # Iter each excel sheet
        # Check out each sheet
        for sheet_name in working_book.sheetnames:
            logging.info(f'Cargando datos: {sheet_name}')

            working_sheet = working_book[sheet_name]

            # Convert sheet to dataframe
            dataframe = pd.DataFrame(working_sheet.values)
            logging.info(f'{dataframe}')
            # Validate categories on local dictionary
            category = find_category(sheet_name)

            if category:
                # Get products category from database
                category_db = execute_sql("SELECT id_categoria FROM categoria WHERE nombre=%s", (category,))

                # Check if dictionary is empty
                if not category_db:
                    logging.error(f'Error cargando datos de la categoria {sheet_name}, '
                                  f'no se encontro la categoria en la base de datos')
                    continue
                else:
                    images_path = save_image_to_s3(working_sheet._images)
                    first_iteration = True
                    for index, row in dataframe.iterrows():
                        if first_iteration:
                            first_iteration = False
                            continue

                        # Clean price and cut data to fill the database
                        row, validation = validate_data(row)

                        if validation:
                            query = """INSERT INTO producto (codigo_ean, titulo, descripcion, precio, url_imagen,
                                    id_categoria, activo) VALUES (%s, %s, %s, %s, %s, %s %s)"""
                            args = (row['CODIGO EAN'], row['REFERENCIA'], row['DESCRIPCION'],
                                    row['PRECIO EN ALMACENES DE CADENA'], images_path[index], category['id_categoria'],
                                    True)
                            execute_sql(query, args)
                        else:
                            return f'Error uploading data: there is an error with file: {index} on {sheet_name} sheet'
            else:
                logging.error(f'Error cargando datos de la categoria {sheet_name}, no se encontro la categoria en los archivos locales')
                continue
    return 'Data uploaded successfully', 200
    # except Exception as e:
    #     return f'Error uploading data: {str(e)}', 500



def find_category(search_value):
    for key, values in categorias.items():
        if isinstance(values, list):
            if search_value in values:
                return key
        else:
            if search_value == values:
                return key
    return None

def save_image_to_s3(images):
    s3_urls  = []

    # iter each image on the images column
    for img in images:
        anchor = img.anchor
        col = anchor._from.col + 1  # openpyxl usa 0-index, sumamos 1 para convertir a 1-index
        logging.info(col)
        if col == 8:  # H1 is at (row=0, col=7)
            #Generate and save a random name for images
            image_name = ''.join(random.choices(string.ascii_letters + string.digits, k=15)) + '.png'

            # Save the image with random name
            image_stream = BytesIO(img._data())

            # Initialize a session using boto3
            s3 = boto3.client('s3', aws_access_key_id=AWS_ACCESS_KEY_ID,
                              aws_secret_access_key=AWS_SECRET_ACCESS_KEY, region_name=AWS_DEFAULT_REGION)
            s3_key = f'imagenes/{image_name}'
            s3.put_object(Bucket=AWS_BUCKET_NAME, Key=s3_key, Body=image_stream,
                          ContentType='image/png')
            s3_url = f'https://{AWS_BUCKET_NAME}.s3.amazonaws.com/{s3_key}'
            s3_urls.append(s3_url)

    return s3_urls


def execute_sql(query, args=None):
    # Connect to the database
    connection = pymysql.connect(host=HOST, port=PORT, user=USER, password=PASSWORD,database=DATABASE,
                                 cursorclass=pymysql.cursors.DictCursor)

    with connection.cursor() as cursor:
        # Execute SQL query
        cursor.execute(query, args)

        # Return category data if exists
        data = cursor.fetchall()
        if data:
            for row in data:
                return row

    # Make commit to save changes
    connection.commit()

    #Close database connection
    if connection:
        connection.close()


def validate_empty_cells(working_book):
    columns_to_check = ['CODIGO EAN', 'REFERENCIA', 'DESCRIPCION', 'PRECIO EN ALMACENES DE CADENA']

    # A list to save results per sheet
    validation_results = []

    # Check out each sheet
    for sheet_name in working_book.sheetnames:
        working_sheet = working_book[sheet_name]

        # Convert sheet to dataframe
        df = pd.DataFrame(working_sheet.values)

        # Get columns name from first row
        new_header = df.iloc[0]
        df = df[1:]
        df.columns = new_header

        # Verify each column for empty data
        empty_cells = {}
        for col in columns_to_check:
            if col in df.columns:
                empty_cells[col] = df[df[col].isnull()].index.tolist()

            # Save results
        if any(empty_cells.values()):
            validation_results.append({
                'HOJA DE EXCEL': sheet_name,
                'Celdas con error': empty_cells
            })

    return validation_results


def validate_data(row):
    # Replace NaN with None
    row['CODIGO EAN'] = row['CODIGO EAN'][:20]
    row['REFERENCIA'] = row['REFERENCIA'][:200]
    row['DESCRIPCION'] = row['DESCRIPCION'][:1000]
    row['PRECIO EN ALMACENES DE CADENA'] = float(row['PRECIO EN ALMACENES DE CADENA'])

    # Validate precio value
    precio_str = str(row['PRECIO EN ALMACENES DE CADENA'])
    # Clean value from non-numeric characters
    precio_str = re.sub(r'[^\d.]', '', precio_str)
    # Turn to float
    row['PRECIO EN ALMACENES DE CADENA'] = float(precio_str)

    return row, True

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8001, debug=True)
